from openpyxl import load_workbook
from model import db, Roster

""" 从excel文件导入学生的花名册
"""
workbook = load_workbook(filename="roster.xlsx")
print(workbook.sheetnames)
sheet = workbook.active
# sheet = workbook['学生基础信息']
# print(sheet)
# cell_obj = sheet.cell(row = 1, column = 1)

# # Print value of cell object
# # using the value attribute
# print(cell_obj.value)

# get max row count
max_row = sheet.max_row
# get max column count
max_column = 2
# max_column=sheet.max_column
# iterate over all cells
# iterate over all rows
for i in range(5, max_row+1):

    # iterate over all columns
    for j in range(1, max_column+1):
        # get particular cell value
        cell_obj = sheet.cell(row=i, column=j)
        # print cell value
        print(cell_obj.value)
        if j == 1:
            code = cell_obj.value
        elif j == 2:
            name = cell_obj.value
        # print(cell_obj.value[1])
    new_roster = Roster(name=name, code=code)
    db.session.add(new_roster)

# print new line
db.session.commit()
print('\n')
workbook.close()
